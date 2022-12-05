from openpyxl import load_workbook


class loadXl():
    
    def __init__(self,filepath):
        self.filepath = filepath
        self.wb = load_workbook(f"{filepath}")
        self.ws = self.wb.active
    def noneFilter(self,data):
        filteredData = tuple(filter(lambda item : item != None,data))
        return filteredData 
    def loadData(self,min_col,max_col,min_row,max_row,valonly = True):
            for precol in self.ws.iter_cols(min_col,max_col,min_row,max_row,valonly):
                return loadXl.noneFilter(self,precol)